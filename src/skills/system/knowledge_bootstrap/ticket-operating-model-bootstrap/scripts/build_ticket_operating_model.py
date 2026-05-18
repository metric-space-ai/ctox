#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


def normalize_header(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text.replace("\n", " ").replace("\r", " "))


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def shorten(text: str, limit: int = 220) -> str:
    text = re.sub(r"\s+", " ", str(text or "").strip())
    return text if len(text) <= limit else text[: limit - 3].rstrip() + "..."


def strip_source_headers(text: str) -> str:
    value = re.sub(r"\s+", " ", str(text or "").strip())
    if not value:
        return value
    patterns = [
        r"(?i)original service case:\s*[^.]+",
        r"(?i)opened in source system:\s*[^.]+",
        r"(?i)channel:\s*[^.]+",
        r"(?i)category:\s*[^.]+",
        r"(?i)subcategory:\s*[^.]+",
        r"(?i)type:\s*[^.]+",
        r"(?i)impact:\s*[^.]+",
        r"(?i)branch/location:\s*[^.]+",
        r"(?i)imported request context:\s*",
    ]
    for pattern in patterns:
        value = re.sub(pattern, " ", value)
    value = re.sub(r"\s+", " ", value).strip(" .;,-")
    return value


def repair_text(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return value
    if "Ã" in value or "â" in value:
        try:
            value = value.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return re.sub(r"\s+", " ", value).strip()


def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    raise TypeError(f"Unsupported type: {type(value)!r}")


def find_column(headers: Iterable[str], *needles: str) -> str | None:
    headers = list(headers)
    lowered = [(h, h.lower()) for h in headers]
    for needle in needles:
        exact = [original for original, lower in lowered if lower == needle]
        if exact:
            return exact[0]
        boundary = [
            original
            for original, lower in lowered
            if re.search(rf"(?<![a-z0-9]){re.escape(needle)}(?![a-z0-9])", lower)
        ]
        if boundary:
            return boundary[0]
        for original, lower in lowered:
            if needle in lower:
                return original
    return None


@dataclass
class TicketRow:
    row_index: int
    values: dict[str, Any]

    def get(self, key: str | None, default: Any = None) -> Any:
        if not key:
            return default
        return self.values.get(key, default)

    @property
    def ticket_id(self) -> str:
        for key in self.values:
            lower = key.lower()
            if "nummer" in lower or "ticket" in lower or "servicefall" in lower:
                value = self.values.get(key)
                if value:
                    return str(value)
        return f"row-{self.row_index}"


def load_rows(path: Path) -> tuple[str, list[TicketRow]]:
    suffix = path.suffix.lower()
    if suffix == ".jsonl":
        return load_jsonl_rows(path)
    if suffix == ".json":
        return load_json_rows(path)
    return load_xlsx_rows(path)


def load_xlsx_rows(path: Path) -> tuple[str, list[TicketRow]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    header = [normalize_header(v) for v in next(iterator)]
    rows: list[TicketRow] = []
    for idx, row in enumerate(iterator, start=2):
        values = {header[i]: row[i] for i in range(len(header))}
        if all(is_blank(v) for v in values.values()):
            continue
        rows.append(TicketRow(idx, values))
    return workbook.sheetnames[0], rows


def load_jsonl_rows(path: Path) -> tuple[str, list[TicketRow]]:
    rows: list[TicketRow] = []
    with path.open("r", encoding="utf-8") as handle:
        for idx, line in enumerate(handle, start=1):
            stripped = line.strip()
            if not stripped:
                continue
            payload = json.loads(stripped)
            if not isinstance(payload, dict):
                continue
            rows.append(
                TicketRow(
                    row_index=idx,
                    values={normalize_header(key): value for key, value in payload.items()},
                )
            )
    return path.name, rows


def load_json_rows(path: Path) -> tuple[str, list[TicketRow]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise SystemExit(f"expected a JSON array of ticket records in {path}")
    rows: list[TicketRow] = []
    for idx, item in enumerate(payload, start=1):
        if not isinstance(item, dict):
            continue
        rows.append(
            TicketRow(
                row_index=idx,
                values={normalize_header(key): value for key, value in item.items()},
            )
        )
    return path.name, rows


def top_tokens(texts: list[str], limit: int = 8) -> list[str]:
    counter: Counter[str] = Counter()
    stop = {
        "und", "der", "die", "das", "mit", "für", "von", "bei", "ein", "eine", "the",
        "oder", "ist", "im", "zu", "wg", "bitte", "auf", "an", "in", "am", "zum"
    }
    for text in texts:
        for token in re.findall(r"[A-Za-zÄÖÜäöüß0-9._/-]{4,}", str(text or "")):
            lower = token.lower()
            if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", token) or re.fullmatch(r"\d{2}\.\d{4}", token):
                continue
            if lower in stop:
                continue
            counter[token] += 1
    return [token for token, _count in counter.most_common(limit)]


def action_patterns(action_texts: list[str]) -> list[str]:
    verbs = [
        "freigeben", "freigegeben", "gesperrt", "entsperrt", "angelegt", "gelöscht",
        "informiert", "weitergeleitet", "geprüft", "erstellt", "aktualisiert",
        "geändert", "installiert", "neu gestartet", "übertragen", "abgeglichen",
    ]
    found = Counter()
    for text in action_texts:
        lower = str(text or "").lower()
        for verb in verbs:
            if verb in lower:
                found[verb] += 1
    return [verb for verb, _count in found.most_common(6)]


def looks_automatic(text: str) -> bool:
    lower = str(text or "").lower()
    automatic_markers = [
        "automatisch",
        "topdesk",
        "wurde automatisch",
        "datum versendet",
        "absender:",
        "betreff:",
        "quarantine summary",
    ]
    return any(marker in lower for marker in automatic_markers)


def curated_note_examples(action_texts: list[str], limit: int = 3) -> list[str]:
    curated = []
    for text in action_texts:
        if not text or looks_automatic(text):
            continue
        cleaned = shorten(strip_source_headers(text), 220)
        if cleaned not in curated:
            curated.append(cleaned)
        if len(curated) >= limit:
            break
    return curated


def curated_request_examples(request_texts: list[str], limit: int = 3) -> list[str]:
    curated = []
    for text in request_texts:
        if not text:
            continue
        cleaned = re.sub(r"(?i)absender:.*?betreff:", "Betreff:", str(text))
        cleaned = shorten(strip_source_headers(cleaned), 220)
        if cleaned not in curated:
            curated.append(cleaned)
        if len(curated) >= limit:
            break
    return curated


def note_style_summary(note_examples: list[str], actions: list[str]) -> str:
    if note_examples:
        if actions:
            return f"Short operational notes with action verbs such as {', '.join(actions[:4])}."
        return "Short operational notes that describe the taken step and its outcome."
    if actions:
        return f"Mostly terse action confirmations around {', '.join(actions[:4])}."
    return "No stable manual note style was recovered; historical notes look mostly automated or sparse."


def infer_note_guidance(mode: str, actions: list[str]) -> str:
    if mode == "access_change":
        guidance = "Write one short internal note that names the affected identity, the access change you checked or performed, and whether the user should retry."
    elif mode == "monitoring_signal":
        guidance = "Write one short internal note that names the affected service or host, the health check or restart you performed, and the current system state."
    elif mode == "integration_issue":
        guidance = "Write one short internal note that names the failing interface or import, the verification step you ran, and whether the next transfer still needs observation."
    elif mode == "notification_or_distribution":
        guidance = "Write one short internal note that records the operational follow-up, the intended recipients or system scope, and whether confirmation is still pending."
    else:
        guidance = "Write one short internal note that records the concrete action taken, the affected object, and the current outcome."
    if actions:
        guidance += f" Prefer the desk's action wording around {', '.join(actions[:3])} when it fits the ticket."
    return guidance


def format_scope_label(request_type: str, category: str, subcategory: str) -> str:
    parts = [repair_text(part) for part in [request_type, category, subcategory] if part and part not in {"unknown", "uncategorized"}]
    return " / ".join(parts) if parts else "diese Ticketfamilie"


def build_operator_summary(
    mode: str,
    request_type: str,
    category: str,
    subcategory: str,
    signals: list[str],
    actions: list[str],
    closure_tendency: float,
) -> str:
    scope = format_scope_label(request_type, category, subcategory)
    if mode == "access_change":
        summary = f"Diese Fälle laufen als {scope}. Zuerst werden Kennung und gewünschte Rechteänderung geklärt, danach die konkrete Sperr-, Entsperr- oder Pflegeaktion festgehalten."
    elif mode == "monitoring_signal":
        summary = f"Diese Fälle laufen als {scope}. Zuerst wird der betroffene Dienst, Host oder Hintergrundjob bestätigt, danach die Gegenmaßnahme oder Gesundheitsprüfung dokumentiert."
    elif mode == "integration_issue":
        summary = f"Diese Fälle laufen als {scope}. Zuerst wird der fehlerhafte Import oder Transfer eingegrenzt, danach der nächste erfolgreiche Lauf oder die Übergabe festgehalten."
    elif mode == "notification_or_distribution":
        summary = f"Diese Fälle laufen als {scope}. Zuerst wird geklärt, ob die Meldung nur informativ ist oder eine konkrete Folgeaktion auslöst."
    else:
        summary = f"Diese Fälle laufen als {scope}. Zuerst wird der betroffene Vorgang sauber eingegrenzt, danach der konkrete Arbeitsschritt knapp dokumentiert."
    if actions:
        summary += f" Wiederkehrende Desk-Verben sind {', '.join(actions[:3])}."
    if closure_tendency >= 0.9:
        summary += " Die Familie wird meist direkt nach der dokumentierten Aktion oder Prüfung geschlossen."
    elif signals:
        summary += f" Starke Wiederholungssignale sind {', '.join(signals[:3])}."
    return summary


def build_note_anchors(mode: str, actions: list[str]) -> list[str]:
    verb_hint = ", ".join(actions[:3]) if actions else "prüfen, dokumentieren, bestätigen"
    if mode == "access_change":
        return [
            "Kurz benennen, welche Kennung oder welches Kurzzeichen betroffen ist.",
            "Die konkrete Sperr-, Entsperr- oder Rechteänderung festhalten.",
            f"Am Ende knapp sagen, ob ein Retry möglich ist. Typische Verben: {verb_hint}.",
        ]
    if mode == "monitoring_signal":
        return [
            "Den betroffenen Host, Dienst oder Job nennen.",
            "Die ausgeführte Prüfung, den Restart oder die Gegenmaßnahme festhalten.",
            f"Den aktuellen Zustand knapp bestätigen. Typische Verben: {verb_hint}.",
        ]
    if mode == "integration_issue":
        return [
            "Die betroffene Schnittstelle oder den Importlauf benennen.",
            "Die geprüfte Ursache oder den Retry knapp festhalten.",
            f"Zum Schluss den sichtbaren Folgezustand nennen. Typische Verben: {verb_hint}.",
        ]
    if mode == "notification_or_distribution":
        return [
            "Nur die operative Folgeaktion notieren, nicht den kompletten Nachrichtentext.",
            "Empfänger, Scope oder Bestätigung kurz nennen.",
            f"Offen lassen, was noch aussteht. Typische Verben: {verb_hint}.",
        ]
    return [
        "Kurz den betroffenen Vorgang oder das Objekt nennen.",
        "Dann den ausgeführten Arbeitsschritt festhalten.",
        f"Zum Schluss den aktuellen Stand oder nächsten Test nennen. Typische Verben: {verb_hint}.",
    ]


def sanitize_decision_text(text: str, limit: int = 220) -> str:
    value = strip_source_headers(repair_text(text))
    value = re.sub(
        r"`(?:triage_focus|handling_steps|decision_support|operator_summary|family_key|historical_examples|close_when|note_guidance|caution_signals)`",
        "",
        value,
    )
    value = re.sub(r"\b(?:sqlite|json dump|parser|yaml|tooling internals|reference commands|ctox ticket)\b", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+", " ", value).strip(" .;,-")
    value = value.replace("`", "")
    if not value:
        return ""
    return shorten(value, limit)


def ticket_title(row: TicketRow, title_col: str | None) -> str:
    return repair_text(str(row.get(title_col) or row.ticket_id))


GENERIC_FAMILY_VALUES = {"", "-", "--", "unknown", "general", "uncategorized", "ticket", "users", "user", "ctox playground", "n/a", "na"}


def extract_structured_fields(*texts: str) -> dict[str, str]:
    combined = "\n".join(repair_text(text) for text in texts if text)
    if not combined:
        return {}
    fields: dict[str, str] = {}
    pattern = re.compile(
        r"(?i)\b([A-ZÄÖÜ][A-Za-zÄÖÜäöüß0-9/_ -]{1,40}):\s*(.+?)(?=(?:\s+[A-ZÄÖÜ][A-Za-zÄÖÜäöüß0-9/_ -]{1,40}:)|$)"
    )
    for label, value in pattern.findall(combined):
        normalized_label = repair_text(label).lower()
        normalized_value = repair_text(value)
        if normalized_label and normalized_value:
            fields[normalized_label] = normalized_value
    return fields


def significant_family_tokens(*texts: str, limit: int = 6) -> list[str]:
    stop = {
        "the", "and", "oder", "und", "with", "from", "into", "nach", "please", "bitte",
        "need", "needs", "required", "wird", "wurde", "sowie", "open", "opened", "ticket",
        "issue", "servicefall", "service", "fall", "case", "user", "konto", "mail", "e-mail",
        "email", "request", "anfrage", "meldung",
    }
    joined = " ".join(repair_text(text) for text in texts if text).lower()
    if not joined:
        return []
    bracketed = re.findall(r"\[([^\]]+)\]", joined)
    tokens = [repair_text(token).lower() for token in bracketed if token.strip()]
    cleaned = re.sub(r"[^a-z0-9äöüß._/-]+", " ", joined)
    for token in cleaned.split():
        token = re.sub(r"\d+", "#", token)
        token = token.strip("#._/-")
        if len(token) < 3 or token in stop:
            continue
        if token not in tokens:
            tokens.append(token)
        if len(tokens) >= limit:
            break
    return tokens


def infer_request_type(raw_value: str, *texts: str) -> str:
    normalized = repair_text(raw_value).lower()
    if normalized and normalized not in GENERIC_FAMILY_VALUES:
        return normalized
    haystack = " ".join(repair_text(text).lower() for text in texts if text)
    if any(marker in haystack for marker in ["critical", "problem", "alert", "warnung", "nagios", "störung", "ausfall"]):
        return "incident"
    if any(marker in haystack for marker in ["sperrung", "entsperr", "passwort", "berechtig", "zugriff", "unlock", "password"]):
        return "access"
    if any(marker in haystack for marker in ["import", "übertrag", "schnittstelle", "hybris", "ftp", "sync", "synchron"]):
        return "integration"
    if any(marker in haystack for marker in ["mail", "e-mail", "email", "sms", "verteiler", "infomail"]):
        return "notification"
    return "ticket"


def infer_category(raw_value: str, *texts: str) -> str:
    normalized = repair_text(raw_value).lower()
    if normalized and normalized not in GENERIC_FAMILY_VALUES:
        return normalized
    tokens = significant_family_tokens(*texts, limit=4)
    if not tokens:
        return "general"
    return " ".join(tokens[:2])


def infer_subcategory(raw_value: str, *texts: str) -> str:
    normalized = repair_text(raw_value).lower()
    if normalized and normalized not in {"", "unknown", "uncategorized"}:
        return normalized
    tokens = significant_family_tokens(*texts, limit=6)
    if len(tokens) >= 4:
        return " ".join(tokens[1:4])
    if len(tokens) >= 2:
        return " ".join(tokens[:3])
    return "uncategorized"


def family_parts(
    row: TicketRow,
    request_type_col: str | None,
    category_col: str | None,
    subcategory_col: str | None,
    title_col: str | None,
    request_col: str | None,
) -> tuple[str, str, str]:
    title = ticket_title(row, title_col)
    request = repair_text(str(row.get(request_col) or ""))
    structured = extract_structured_fields(title, request)
    request_type = infer_request_type(
        structured.get("type", str(row.get(request_type_col) or "")),
        title,
        request,
    )
    category = infer_category(
        structured.get("category", str(row.get(category_col) or "")),
        title,
        request,
    )
    subcategory = infer_subcategory(
        structured.get("subcategory", str(row.get(subcategory_col) or "")),
        title,
        request,
    )
    return request_type, category, subcategory


def example_card(row: TicketRow, title_col: str | None, why: str) -> dict[str, Any]:
    return {
        "ticket_id": row.ticket_id,
        "title": shorten(ticket_title(row, title_col), 140),
        "why": why,
    }


def family_key(
    row: TicketRow,
    request_type_col: str | None,
    category_col: str | None,
    subcategory_col: str | None,
    title_col: str | None,
    request_col: str | None,
) -> str:
    parts = family_parts(row, request_type_col, category_col, subcategory_col, title_col, request_col)
    return " :: ".join(parts)


def top_phrases(texts: list[str], limit: int = 8) -> list[str]:
    counter: Counter[str] = Counter()
    for text in texts:
        cleaned = repair_text(text).lower()
        cleaned = re.sub(r"[^a-z0-9äöüß._/-]+", " ", cleaned)
        tokens = [token for token in cleaned.split() if len(token) >= 4]
        for i in range(len(tokens) - 1):
            phrase = f"{tokens[i]} {tokens[i + 1]}"
            if "datum" in phrase or "betreff" in phrase or "versendet" in phrase:
                continue
            if re.search(r"\b\d{2}\.\d{2}\.\d{4}\b", phrase) or re.search(r"\b\d{2}\.\d{4}\b", phrase):
                continue
            counter[phrase] += 1
    return [phrase for phrase, _count in counter.most_common(limit)]


def infer_family_mode(signals: list[str], actions: list[str], request_type: str, category: str, subcategory: str) -> str:
    haystack = " ".join([request_type, category, subcategory, *signals, *actions]).lower()
    if any(marker in haystack for marker in ["passwort", "berecht", "kurzzeichen", "bediener", "entsperrt", "gesperrt"]):
        return "access_change"
    if any(marker in haystack for marker in ["nagios", "alert", "critical", "systemmeldung", "hintergrundprogramm", "aix"]):
        return "monitoring_signal"
    if any(marker in haystack for marker in ["hybris", "übertragung", "schnittstelle", "import", "maufto"]):
        return "integration_issue"
    if any(marker in haystack for marker in ["sms", "mail", "infomail"]):
        return "notification_or_distribution"
    return "service_request"


def infer_handling_steps(
    mode: str,
    request_type: str,
    category: str,
    subcategory: str,
    signals: list[str],
    actions: list[str],
    closure_tendency: float,
) -> list[str]:
    request_scope = " / ".join(part for part in [request_type, category, subcategory] if part and part != "unknown")
    steps: list[str] = [f"Match the ticket against the repeated family scope `{request_scope}` and confirm the affected service, user, or system from the subject and request text."]
    if mode == "access_change":
        steps.extend(
            [
                "Verify which user, short code, or account is affected and whether this is a lock, unlock, password change, or entitlement change.",
                "Check whether the request is an operational follow-up to an automated signal or a direct access request from a colleague or store.",
                "Perform the access change only after the target identity and requested scope are unambiguous.",
                "Leave a short internal note that records the identity, the change made, and whether the user should retry immediately.",
            ]
        )
    elif mode == "monitoring_signal":
        steps.extend(
            [
                "Identify the referenced host, service, or background process from the alert wording before taking action.",
                "Check whether a restart, re-run, or simple verification is historically the normal first response for this family.",
                "Only close once the monitoring signal cleared or the affected process is confirmed healthy again.",
            ]
        )
    elif mode == "integration_issue":
        steps.extend(
            [
                "Confirm which interface, import, or background transfer failed and whether the failure is single-shot or recurring.",
                "Inspect the failing run or payload before retrying or escalating.",
                "Close only after the next successful transfer, import, or synchronization run is visible.",
            ]
        )
    elif mode == "notification_or_distribution":
        steps.extend(
            [
                "Determine whether the ticket is informational only or requires an operational reaction.",
                "If action is needed, capture the concrete follow-up rather than restating the raw mail content.",
                "Close once distribution, verification, or acknowledgement is complete.",
            ]
        )
    else:
        steps.extend(
            [
                "Confirm the request scope and missing information before taking an operational step.",
                "Use the repeated historical action wording as a cue for the likely next move.",
                "Close only once the requested change or verification is explicitly done.",
            ]
        )
    if actions:
        steps.append(f"Historical action wording to mirror carefully: {', '.join(actions[:4])}.")
    if closure_tendency >= 0.9:
        steps.append("This family historically closes quickly once the concrete action or verification is documented.")
    if any(token.lower() in {"datum", "betreff", "versendet"} for token in signals):
        steps.append("Ignore mail-wrapper noise when reading the request; focus on the operational noun phrases instead.")
    return steps[:7]


def infer_close_when(mode: str, closure_tendency: float) -> str:
    if mode == "access_change":
        return "Close when the identity change is documented and the affected user or store can retry with the intended access."
    if mode == "monitoring_signal":
        return "Close when the alert condition is gone or the monitored service is verified healthy again."
    if mode == "integration_issue":
        return "Close when the next interface or transfer run succeeds and the failure no longer reproduces."
    if mode == "notification_or_distribution":
        return "Close when the intended recipients, acknowledgement, or follow-up verification are complete."
    if closure_tendency >= 0.9:
        return "Close once the concrete handling step is done and briefly documented."
    return "Close only after the requested outcome or verification is explicit in the working note."


def infer_caution_signals(mode: str, signals: list[str], actions: list[str]) -> list[str]:
    cautions = []
    if mode == "access_change":
        cautions.append("Do not change access when the affected identity or requested scope is still ambiguous.")
    if mode == "monitoring_signal":
        cautions.append("Do not treat every monitoring mail as action-free; verify whether this family usually triggers a restart or operator check.")
    if mode == "integration_issue":
        cautions.append("Do not close on the first symptom description alone; look for evidence of a successful rerun.")
    if any(signal.lower() in {"datum", "betreff", "versendet"} for signal in signals):
        cautions.append("Mail-wrapper text is noisy in this family and can hide the real issue behind repeated transport metadata.")
    if not actions:
        cautions.append("Historical action notes are sparse, so example tickets matter more than verb extraction here.")
    return cautions[:4]


def build_decision_support(
    request_type: str,
    category: str,
    subcategory: str,
    signals: list[str],
    phrases: list[str],
    actions: list[str],
    closure_tendency: float,
    note_examples: list[str],
) -> dict[str, Any]:
    mode = infer_family_mode(signals, actions, request_type, category, subcategory)
    return {
        "mode": mode,
        "operator_summary": sanitize_decision_text(
            build_operator_summary(mode, request_type, category, subcategory, signals, actions, closure_tendency)
        ),
        "triage_focus": [sanitize_decision_text(item, 120) for item in (phrases[:5] or signals[:5]) if sanitize_decision_text(item, 120)],
        "handling_steps": infer_handling_steps(
            mode,
            request_type,
            category,
            subcategory,
            signals,
            actions,
            closure_tendency,
        ),
        "close_when": sanitize_decision_text(infer_close_when(mode, closure_tendency), 180),
        "caution_signals": [
            sanitize_decision_text(item, 180)
            for item in infer_caution_signals(mode, signals, actions)
            if sanitize_decision_text(item, 180)
        ],
        "note_guidance": sanitize_decision_text(infer_note_guidance(mode, actions), 220),
        "note_anchors": [
            sanitize_decision_text(item, 180)
            for item in build_note_anchors(mode, actions)
            if sanitize_decision_text(item, 180)
        ],
    }


def openai_output_text(payload: dict[str, Any]) -> str:
    if isinstance(payload.get("output_text"), str) and payload["output_text"].strip():
        return payload["output_text"].strip()
    for item in payload.get("output", []):
        if item.get("type") != "message":
            continue
        for content in item.get("content", []):
            if content.get("type") == "output_text" and content.get("text"):
                return str(content["text"]).strip()
    return ""


def normalize_string_list(value: Any, limit: int) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith("[") and stripped.endswith("]"):
            try:
                parsed = json.loads(stripped)
                if isinstance(parsed, list):
                    return [repair_text(str(item)) for item in parsed if str(item).strip()][:limit]
            except json.JSONDecodeError:
                pass
        return [repair_text(stripped)] if stripped else []
    if isinstance(value, list):
        return [repair_text(str(item)) for item in value if str(item).strip()][:limit]
    return [repair_text(str(value))][:limit]


def normalize_string_field(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        values = [repair_text(str(item)) for item in value if str(item).strip()]
        return "; ".join(values)
    return repair_text(str(value))


def refine_decision_support_openai(
    family_key: str,
    evidence: dict[str, Any],
    api_key: str,
    model: str,
    base_url: str,
) -> dict[str, Any] | None:
    prompt = {
        "family_key": family_key,
        "signals": evidence["signals"],
        "common_phrases": evidence["common_phrases"],
        "actions_seen": evidence["actions_seen"],
        "working_pattern": evidence["working_pattern"],
        "closure_tendency": evidence["closure_tendency"],
        "request_examples": evidence["request_examples"],
        "note_examples": evidence["note_examples"],
        "historical_example_titles": evidence["historical_example_titles"],
    }
    body = {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "You are refining a helpdesk operating-model playbook from historical ticket evidence. "
                            "Return strict JSON only. Do not invent tools, databases, or policies. "
                            "Summarize how operators actually handle this ticket family. Keep the language compact and operational."
                        ),
                    }
                ],
            },
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": (
                            "Using the evidence below, produce JSON with keys "
                            "`operator_summary`, `triage_focus`, `handling_steps`, `close_when`, `caution_signals`, `note_guidance`. "
                            "`triage_focus`, `handling_steps`, and `caution_signals` must be arrays of short strings. "
                            "Base everything only on the evidence.\n\n"
                            + json.dumps(prompt, ensure_ascii=False, indent=2)
                        ),
                    }
                ],
            },
        ],
        "text": {"format": {"type": "json_object"}},
    }
    request = urllib.request.Request(
        f"{base_url.rstrip('/')}/v1/responses",
        data=json.dumps(body).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError):
        return None
    text = openai_output_text(payload)
    if not text:
        return None
    try:
        refined = json.loads(text)
    except json.JSONDecodeError:
        return None
    if not isinstance(refined, dict):
        return None
    return refined


def build_operating_model(
    sheet_name: str,
    rows: list[TicketRow],
    top_families: int,
    min_family_size: int,
    openai_model: str | None = None,
    openai_base_url: str | None = None,
    openai_api_key: str | None = None,
    openai_refine_limit: int = 0,
) -> dict[str, Any]:
    headers = rows[0].values.keys()
    title_col = find_column(headers, "kurzbeschreibung", "details", "title", "subject")
    request_type_col = find_column(
        headers,
        "art 'servicefall'",
        "art servicefall",
        "request_type",
        "ticket_type",
        "type",
    )
    category_col = find_column(headers, "kategorie", "category", "group", "group_name", "queue")
    subcategory_col = find_column(
        headers,
        "unterkategorie",
        "subcategory",
        "sub_type",
        "label",
        "primary_label",
    )
    channel_col = find_column(headers, "eingangsart", "channel", "source_channel")
    state_col = find_column(headers, "ausführung", "state", "remote_status", "status")
    impact_col = find_column(headers, "auswirkung", "impact", "priority")
    reporter_col = find_column(headers, "name anmelder", "requester", "reporter", "customer")
    owner_col = find_column(headers, "owner", "owner_name", "assignee", "agent", "user")
    group_col = find_column(headers, "group", "group_name", "queue")
    request_col = find_column(headers, "anfrage", "request_text", "body_text", "body", "request")
    action_col = find_column(headers, "aktion", "action_text", "latest_action", "latest_reply")

    families: dict[str, list[TicketRow]] = defaultdict(list)
    for row in rows:
        families[family_key(row, request_type_col, category_col, subcategory_col, title_col, request_col)].append(row)

    ranked_families = sorted(families.items(), key=lambda item: len(item[1]), reverse=True)
    selected = [(key, family_rows) for key, family_rows in ranked_families if len(family_rows) >= min_family_size][:top_families]

    operating_families = []
    family_playbooks = []
    state_norms = []
    note_style_refs = []
    retrieval_cards = []

    for index, (key, family_rows) in enumerate(selected, start=1):
        request_type, category, subcategory = key.split(" :: ")
        channels = Counter(str(row.get(channel_col) or "unknown").strip() for row in family_rows if not is_blank(row.get(channel_col)))
        states = Counter(str(row.get(state_col) or "unknown").strip() for row in family_rows if not is_blank(row.get(state_col)))
        impacts = Counter(str(row.get(impact_col) or "unknown").strip() for row in family_rows if not is_blank(row.get(impact_col)))
        reporters = Counter(str(row.get(reporter_col) or "unknown").strip() for row in family_rows if not is_blank(row.get(reporter_col)))
        owners = Counter(str(row.get(owner_col) or "unknown").strip() for row in family_rows if not is_blank(row.get(owner_col)))
        groups = Counter(str(row.get(group_col) or "unknown").strip() for row in family_rows if not is_blank(row.get(group_col)))

        request_texts = [repair_text(str(row.get(request_col) or "")) for row in family_rows[:200]]
        action_texts = [repair_text(str(row.get(action_col) or "")) for row in family_rows[:200]]
        title_texts = [ticket_title(row, title_col) for row in family_rows[:200]]
        signals = top_tokens(title_texts + request_texts, limit=8)
        phrases = top_phrases(title_texts + request_texts, limit=8)
        actions = action_patterns(action_texts)
        manual_note_examples = curated_note_examples(action_texts, limit=3)
        curated_requests = curated_request_examples(request_texts, limit=3)
        dominant_state_names = [state for state, _count in states.most_common(5)]
        closure_tendency = sum(
            count for state, count in states.items() if state.lower() in {"geschlossen", "fertig"}
        ) / max(1, len(family_rows))

        canonical_examples = [example_card(row, title_col, "Representative historical case for this family.") for row in family_rows[:2]]
        common_examples = [example_card(row, title_col, "Common historical case in this family.") for row in family_rows[2:5]]
        closure_examples = [
            example_card(row, title_col, "Family example that reached a finished or closed state.")
            for row in family_rows
            if str(row.get(state_col) or "").strip().lower() in {"geschlossen", "fertig"}
        ][:3]

        working_pattern = (
            f"Usually enters through {channels.most_common(1)[0][0] if channels else 'unknown'}, "
            f"moves through {', '.join(dominant_state_names[:3]) or 'unknown states'}, "
            f"and often ends in {', '.join(state for state in dominant_state_names if state.lower() in {'geschlossen', 'fertig'}) or 'no stable closure'}."
        )
        decision_support = build_decision_support(
            request_type,
            category,
            subcategory,
            signals,
            phrases,
            actions,
            closure_tendency,
            manual_note_examples,
        )
        evidence = {
            "signals": signals,
            "common_phrases": phrases,
            "actions_seen": actions,
            "working_pattern": working_pattern,
            "closure_tendency": closure_tendency,
            "request_examples": curated_requests,
            "note_examples": manual_note_examples,
            "historical_example_titles": [example["title"] for example in canonical_examples + common_examples[:2]],
        }
        if openai_model and openai_api_key and index <= openai_refine_limit:
            refined = refine_decision_support_openai(
                key,
                evidence,
                openai_api_key,
                openai_model,
                openai_base_url or "https://api.openai.com",
            )
            if refined:
                decision_support.update(
                    {
                        "operator_summary": sanitize_decision_text(
                            normalize_string_field(refined.get("operator_summary") or decision_support["operator_summary"])
                        ),
                        "triage_focus": [
                            sanitize_decision_text(item, 120)
                            for item in normalize_string_list(refined.get("triage_focus") or decision_support["triage_focus"], 6)
                            if sanitize_decision_text(item, 120)
                        ],
                        "handling_steps": [
                            sanitize_decision_text(item, 220)
                            for item in normalize_string_list(refined.get("handling_steps") or decision_support["handling_steps"], 8)
                            if sanitize_decision_text(item, 220)
                        ],
                        "close_when": sanitize_decision_text(
                            normalize_string_field(refined.get("close_when") or decision_support["close_when"]),
                            180,
                        ),
                        "caution_signals": [
                            sanitize_decision_text(item, 180)
                            for item in normalize_string_list(refined.get("caution_signals") or decision_support["caution_signals"], 5)
                            if sanitize_decision_text(item, 180)
                        ],
                        "note_guidance": sanitize_decision_text(
                            normalize_string_field(refined.get("note_guidance") or decision_support["note_guidance"]),
                            220,
                        ),
                    }
                )

        operating_families.append(
            {
                "family_key": key,
                "request_type": request_type,
                "category": category,
                "subcategory": subcategory,
                "row_count": len(family_rows),
                "dominant_channels": channels.most_common(3),
                "dominant_states": states.most_common(5),
                "dominant_impacts": impacts.most_common(3),
                "dominant_reporters": reporters.most_common(5),
                "dominant_groups": groups.most_common(5),
                "dominant_owners": owners.most_common(5),
                "likely_source_systems": reporters.most_common(3),
            }
        )
        family_playbooks.append(
            {
                "family_key": key,
                "rank": index,
                "signals": {
                    "request_type": request_type,
                    "category": category,
                    "subcategory": subcategory,
                    "token_signals": signals,
                    "common_phrases": phrases,
                },
                "usual_handling": {
                    "row_count": len(family_rows),
                    "dominant_channels": channels.most_common(3),
                    "dominant_states": states.most_common(5),
                    "dominant_groups": groups.most_common(5),
                    "dominant_owners": owners.most_common(5),
                    "actions_seen": actions,
                    "closure_tendency": closure_tendency,
                    "working_pattern": working_pattern,
                },
                "decision_support": decision_support,
                "note_style": {
                    "summary": note_style_summary(manual_note_examples, actions),
                    "manual_note_examples": manual_note_examples,
                },
                "historical_examples": {
                    "canonical": canonical_examples,
                    "common": common_examples,
                    "closure": closure_examples,
                },
                "likely_sources": reporters.most_common(5),
                "likely_groups": groups.most_common(5),
                "likely_owners": owners.most_common(5),
            }
        )
        state_norms.append(
            {
                "family_key": key,
                "dominant_states": states.most_common(5),
                "dominant_closure_states": [(state, count) for state, count in states.most_common() if state.lower() in {"geschlossen", "fertig"}][:3],
                "closure_tendency": closure_tendency,
            }
        )
        note_style_refs.append(
            {
                "family_key": key,
                "note_examples": manual_note_examples,
                "request_examples": curated_requests,
                "note_style_summary": note_style_summary(manual_note_examples, actions),
            }
        )

        family_card = {
            "card_id": f"family:{index}",
            "card_type": "family_playbook",
            "family_key": key,
            "request_type": request_type,
            "category": category,
            "subcategory": subcategory,
            "text": " | ".join(
                [
                    key,
                    "signals " + ", ".join(signals[:5]),
                    "phrases " + ", ".join(phrases[:4]),
                    "actions " + ", ".join(actions[:5]),
                    "states " + ", ".join(state for state, _count in states.most_common(3)),
                    "working " + (
                        f"channel {channels.most_common(1)[0][0]}" if channels else "channel unknown"
                    ),
                    "examples " + " ; ".join(example["title"] for example in canonical_examples[:2]),
                ]
            ),
        }
        retrieval_cards.append(family_card)
        for example in canonical_examples + common_examples:
            retrieval_cards.append(
                {
                    "card_id": f"example:{example['ticket_id']}",
                    "card_type": "historical_example",
                    "family_key": key,
                    "request_type": request_type,
                    "category": category,
                    "subcategory": subcategory,
                    "text": f"{key} | {example['title']} | {' ; '.join(curated_requests[:1])} | actions {', '.join(actions[:3])}",
                    "ticket_id": example["ticket_id"],
                    "title": example["title"],
                }
            )

    return {
        "source_scope": {"sheet_name": sheet_name, "row_count": len(rows)},
        "fields": {
            "request_type": request_type_col,
            "category": category_col,
            "subcategory": subcategory_col,
            "channel": channel_col,
            "state": state_col,
            "impact": impact_col,
            "reporter": reporter_col,
            "request_text": request_col,
            "action_text": action_col,
        },
        "operating_families": operating_families,
        "family_playbooks": family_playbooks,
        "state_transition_norms": state_norms,
        "note_style_refs": note_style_refs,
        "retrieval_cards": retrieval_cards,
    }


def maybe_build_vectors(cards: list[dict[str, Any]], provider: str | None, model_name: str | None) -> list[list[float]] | None:
    if not provider or not model_name:
        return None
    if provider != "sentence-transformers":
        raise ValueError(f"Unsupported embedding provider: {provider}")
    from sentence_transformers import SentenceTransformer
    import numpy as np
    import torch

    model = SentenceTransformer(model_name, device="cuda" if torch.cuda.is_available() else "cpu")
    texts = [card["text"] for card in cards]
    vectors = model.encode(texts, batch_size=32, show_progress_bar=False, normalize_embeddings=True)
    return np.asarray(vectors).tolist()


def render_operating_model_markdown(model: dict[str, Any]) -> str:
    lines = ["# Ticket Operating Model", ""]
    scope = model["source_scope"]
    lines.append(f"- Sheet: `{scope['sheet_name']}`")
    lines.append(f"- Rows: `{scope['row_count']}`")
    lines.append("")
    lines.append("## How This Desk Appears To Work")
    lines.append("")
    for playbook in model["family_playbooks"][:10]:
        handling = playbook["usual_handling"]
        note_style = playbook["note_style"]
        lines.append(f"### {playbook['family_key']}")
        lines.append("")
        lines.append(f"- Signals: {', '.join(playbook['signals']['token_signals'][:6])}")
        lines.append(f"- Dominant states: {', '.join(state for state, _count in handling['dominant_states'][:4])}")
        lines.append(f"- Dominant channels: {', '.join(channel for channel, _count in handling['dominant_channels'][:3])}")
        lines.append(f"- Repeated actions: {', '.join(handling['actions_seen'][:5]) or 'none observed'}")
        lines.append(f"- Working pattern: {handling['working_pattern']}")
        lines.append(f"- Operator summary: {playbook['decision_support']['operator_summary']}")
        lines.append(f"- First checks: {'; '.join(playbook['decision_support']['triage_focus'][:4])}")
        lines.append(f"- Handling steps: {'; '.join(playbook['decision_support']['handling_steps'][:4])}")
        lines.append(f"- Close when: {playbook['decision_support']['close_when']}")
        lines.append(f"- Note style: {note_style['summary']}")
        lines.append(f"- Closure tendency: {handling['closure_tendency']:.2f}")
        examples = ", ".join(example["ticket_id"] for example in playbook["historical_examples"]["canonical"][:2])
        lines.append(f"- Canonical examples: {examples}")
        lines.append("")
    return "\n".join(lines) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a ticket operating model from a historical ticket dataset.")
    parser.add_argument("--input")
    parser.add_argument("--input-xlsx")
    parser.add_argument("--input-jsonl")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--top-families", type=int, default=20)
    parser.add_argument("--min-family-size", type=int, default=20)
    parser.add_argument("--embedding-provider")
    parser.add_argument("--embedding-model")
    parser.add_argument("--openai-model")
    parser.add_argument("--openai-base-url", default=os.getenv("OPENAI_BASE_URL", "https://api.openai.com"))
    parser.add_argument("--openai-api-key-env", default="OPENAI_API_KEY")
    parser.add_argument("--openai-refine-limit", type=int, default=0)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = args.input or args.input_xlsx or args.input_jsonl
    if not input_path:
        raise SystemExit("one of --input, --input-xlsx, or --input-jsonl is required")
    sheet_name, rows = load_rows(Path(input_path))
    model = build_operating_model(
        sheet_name,
        rows,
        args.top_families,
        args.min_family_size,
        openai_model=args.openai_model,
        openai_base_url=args.openai_base_url,
        openai_api_key=os.getenv(args.openai_api_key_env) if args.openai_model else None,
        openai_refine_limit=args.openai_refine_limit,
    )
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    (output_dir / "operating_families.json").write_text(
        json.dumps(model["operating_families"], ensure_ascii=False, indent=2, default=json_default) + "\n",
        encoding="utf-8",
    )
    (output_dir / "family_playbooks.json").write_text(
        json.dumps(model["family_playbooks"], ensure_ascii=False, indent=2, default=json_default) + "\n",
        encoding="utf-8",
    )
    (output_dir / "state_transition_norms.json").write_text(
        json.dumps(model["state_transition_norms"], ensure_ascii=False, indent=2, default=json_default) + "\n",
        encoding="utf-8",
    )
    (output_dir / "note_style_refs.json").write_text(
        json.dumps(model["note_style_refs"], ensure_ascii=False, indent=2, default=json_default) + "\n",
        encoding="utf-8",
    )
    with (output_dir / "retrieval_index.jsonl").open("w", encoding="utf-8") as handle:
        for card in model["retrieval_cards"]:
            handle.write(json.dumps(card, ensure_ascii=False) + "\n")

    vectors = maybe_build_vectors(model["retrieval_cards"], args.embedding_provider, args.embedding_model)
    if vectors is not None:
        import numpy as np

        np.save(output_dir / "retrieval_vectors.npy", np.asarray(vectors))

    (output_dir / "operating_model.md").write_text(render_operating_model_markdown(model), encoding="utf-8")

    print(
        json.dumps(
            {
                "sheet_name": sheet_name,
                "row_count": len(rows),
                "family_count": len(model["operating_families"]),
                "retrieval_card_count": len(model["retrieval_cards"]),
                "vectors_built": vectors is not None,
                "output_dir": str(output_dir),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
