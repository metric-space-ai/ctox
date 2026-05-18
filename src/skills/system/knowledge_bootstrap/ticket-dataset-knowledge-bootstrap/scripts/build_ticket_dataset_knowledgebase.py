#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

def normalize_header(value: Any) -> str:
    text = str(value or "").strip()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def shorten(text: str, limit: int = 240) -> str:
    text = re.sub(r"\s+", " ", text.strip())
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def json_default(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    raise TypeError(f"Unsupported type: {type(value)!r}")


@dataclass
class TicketRow:
    row_index: int
    values: dict[str, Any]

    def get(self, key: str, default: Any = None) -> Any:
        return self.values.get(key, default)

    @property
    def ticket_id(self) -> str:
        for key in self.values:
            lower = key.lower()
            if "nummer" in lower or "ticket" in lower or "servicefall" in lower:
                value = self.values.get(key)
                if value:
                    return str(value)
        return f"row-{self.row_index}"


def find_column(headers: Iterable[str], *needles: str) -> str | None:
    headers = list(headers)
    lowered = [(h, h.lower()) for h in headers]
    for needle in needles:
        exact_matches = [original for original, lower in lowered if lower == needle]
        if exact_matches:
            return exact_matches[0]
        boundary_matches = [
            original
            for original, lower in lowered
            if re.search(rf"(?<![a-z0-9]){re.escape(needle)}(?![a-z0-9])", lower)
        ]
        if boundary_matches:
            return boundary_matches[0]
        for original, lower in lowered:
            if needle in lower:
                return original
    return None


def load_ticket_rows(path: Path) -> tuple[str, list[TicketRow]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    header = [normalize_header(v) for v in next(iterator)]
    rows: list[TicketRow] = []
    for idx, row in enumerate(iterator, start=2):
        values = {header[i]: row[i] for i in range(len(header))}
        if all(is_blank(v) for v in values.values()):
            continue
        rows.append(TicketRow(row_index=idx, values=values))
    return wb.sheetnames[0], rows


def classify_column(name: str, values: list[Any]) -> dict[str, Any]:
    non_null = [v for v in values if not is_blank(v)]
    unique = {str(v).strip() for v in non_null}
    fill_ratio = len(non_null) / max(1, len(values))
    lower = name.lower()

    role = "unknown"
    if any(isinstance(v, (datetime, date)) for v in non_null) or "datum" in lower or "termin" in lower:
        role = "time"
    elif "nummer" in lower or ("servicefall" in lower and len(unique) > 0.7 * len(non_null)):
        role = "id"
    elif "beschreibung" in lower or "anfrage" in lower or "aktion" in lower or "details" in lower:
        role = "free_text"
    elif "kategorie" in lower or "art" in lower or "eingangsart" in lower:
        role = "category_or_tag"
    elif "ausführung" in lower or "status" in lower:
        role = "enum_state"
    elif "name" in lower or "filiale" in lower:
        role = "owner_or_team"
    elif len(unique) <= max(40, len(non_null) * 0.02):
        role = "enum_state" if any(x in lower for x in ["status", "ausführung", "reaktion"]) else "category_or_tag"
    elif all(isinstance(v, (int, float)) for v in non_null[:20] if v is not None):
        role = "metric"

    top_values = Counter(str(v).strip() for v in non_null).most_common(8)
    return {
        "column": name,
        "role": role,
        "fill_ratio": round(fill_ratio, 4),
        "non_null_count": len(non_null),
        "distinct_count": len(unique),
        "top_values": [{"value": key, "count": count} for key, count in top_values],
    }


def build_source_profile(sheet_name: str, rows: list[TicketRow]) -> dict[str, Any]:
    headers = list(rows[0].values.keys())
    by_column = {h: [row.get(h) for row in rows] for h in headers}
    columns = [classify_column(h, values) for h, values in by_column.items()]
    return {
        "source_scope": {
            "sheet_name": sheet_name,
            "row_count": len(rows),
            "column_count": len(headers),
        },
        "row_unit_description": "One row represents one historical service ticket / service case.",
        "tables_or_sheets": [sheet_name],
        "key_columns": [c["column"] for c in columns if c["role"] == "id"],
        "enum_like_columns": [c["column"] for c in columns if c["role"] in {"enum_state", "category_or_tag"}],
        "free_text_columns": [c["column"] for c in columns if c["role"] == "free_text"],
        "ownership_columns": [c["column"] for c in columns if c["role"] == "owner_or_team"],
        "reference_columns": [c["column"] for c in columns if c["role"] == "reference"],
        "columns": columns,
        "ambiguities": [
            "Name-oriented columns may represent reporter systems, functional mailboxes, or human requesters.",
            "Action text may encode closure automation, manual work notes, or escalation history in the same field.",
        ],
    }


def example_for_row(row: TicketRow, title_column: str | None) -> dict[str, Any]:
    title = str(row.get(title_column) or row.ticket_id) if title_column else row.ticket_id
    return {"ticket_id": row.ticket_id, "title": shorten(title, 140)}


def promote_structured_taxonomies(rows: list[TicketRow]) -> list[dict[str, Any]]:
    headers = rows[0].values.keys()
    title_column = find_column(headers, "kurzbeschreibung", "details")
    candidate_fields = [
        ("channel_family", find_column(headers, "eingangsart")),
        ("request_type", find_column(headers, "art 'servicefall'", "art servicefall")),
        ("service_category", find_column(headers, "kategorie")),
        ("subcategory", find_column(headers, "unterkategorie")),
        ("impact_scope", find_column(headers, "auswirkung")),
        ("workflow_state", find_column(headers, "ausführung")),
        ("reporter_family", find_column(headers, "name anmelder")),
        ("site_scope", find_column(headers, "filiale")),
    ]

    taxonomies: list[dict[str, Any]] = []
    for dimension_name, field in candidate_fields:
        if not field:
            continue
        buckets: dict[str, list[TicketRow]] = defaultdict(list)
        for row in rows:
            value = row.get(field)
            if is_blank(value):
                continue
            buckets[str(value).strip()].append(row)
        if len(buckets) < 2:
            continue
        promoted = []
        for value, bucket_rows in sorted(buckets.items(), key=lambda item: len(item[1]), reverse=True)[:12]:
            if len(bucket_rows) < 5:
                continue
            promoted.append(
                {
                    "bucket_name": value,
                    "definition": f"Rows where `{field}` resolves to `{value}`.",
                    "row_count_estimate": len(bucket_rows),
                    "confidence": round(min(0.99, 0.6 + min(len(bucket_rows), 3000) / 10000), 2),
                    "canonical_examples": [example_for_row(r, title_column) for r in bucket_rows[:2]],
                    "common_examples": [example_for_row(r, title_column) for r in bucket_rows[2:5]],
                    "edge_examples": [],
                    "inclusion_signals": [f"{field}={value}"],
                    "exclusion_signals": [f"{field}!={value}"],
                }
            )
        if promoted:
            taxonomies.append(
                {
                    "dimension_name": dimension_name,
                    "purpose": f"Structured taxonomy promoted from `{field}`.",
                    "inference_basis": "direct_structured_column",
                    "source_fields": [field],
                    "bucket_count": len(promoted),
                    "stability_note": "Promoted because the column behaves like a stable categorical control field.",
                    "state": "promoted",
                    "buckets": promoted,
                }
            )

    subcategory = find_column(headers, "unterkategorie")
    category = find_column(headers, "kategorie")
    if category and subcategory:
        combo: dict[str, list[TicketRow]] = defaultdict(list)
        for row in rows:
            left = row.get(category)
            right = row.get(subcategory)
            if is_blank(left) and is_blank(right):
                continue
            combo[f"{left or 'unknown'} > {right or 'unknown'}"].append(row)
        buckets = []
        for value, bucket_rows in sorted(combo.items(), key=lambda item: len(item[1]), reverse=True)[:15]:
            if len(bucket_rows) < 8:
                continue
            buckets.append(
                {
                    "bucket_name": value,
                    "definition": f"Rows where category/subcategory co-occur as `{value}`.",
                    "row_count_estimate": len(bucket_rows),
                    "confidence": 0.83,
                    "canonical_examples": [example_for_row(r, title_column) for r in bucket_rows[:2]],
                    "common_examples": [example_for_row(r, title_column) for r in bucket_rows[2:5]],
                    "edge_examples": [],
                    "inclusion_signals": [value],
                    "exclusion_signals": [],
                }
            )
        if buckets:
            taxonomies.append(
                {
                    "dimension_name": "category_tree",
                    "purpose": "Operational family built from category and subcategory together.",
                    "inference_basis": "paired_structured_columns",
                    "source_fields": [category, subcategory],
                    "bucket_count": len(buckets),
                    "stability_note": "Promoted because category and subcategory together form clearer operational families than either field alone.",
                    "state": "promoted",
                    "buckets": buckets,
                }
            )
    return taxonomies


def build_semantic_text(row: TicketRow, headers: Iterable[str]) -> str:
    parts = []
    for field in [
        find_column(headers, "kurzbeschreibung", "details"),
        find_column(headers, "unterkategorie"),
        find_column(headers, "kategorie"),
        find_column(headers, "name anmelder"),
        find_column(headers, "art 'servicefall'", "art servicefall"),
        find_column(headers, "anfrage"),
        find_column(headers, "aktion"),
    ]:
        if not field:
            continue
        value = row.get(field)
        if is_blank(value):
            continue
        text = str(value).strip()
        if "anfrage" in field.lower() or "aktion" in field.lower():
            text = shorten(text, 800)
        parts.append(f"{field}: {text}")
    return "\n".join(parts)


def sample_rows_for_semantics(rows: list[TicketRow], max_rows: int) -> list[TicketRow]:
    if len(rows) <= max_rows:
        return rows
    headers = rows[0].values.keys()
    stratify_field = find_column(headers, "unterkategorie", "kategorie", "eingangsart")
    if not stratify_field:
        step = max(1, len(rows) // max_rows)
        return rows[::step][:max_rows]
    buckets: dict[str, list[TicketRow]] = defaultdict(list)
    for row in rows:
        buckets[str(row.get(stratify_field) or "unknown")].append(row)
    sampled: list[TicketRow] = []
    for bucket_rows in buckets.values():
        take = max(3, int(round(max_rows * len(bucket_rows) / len(rows))))
        sampled.extend(bucket_rows[:take])
    return sampled[:max_rows]


def load_sentence_transformer(model_name: str):
    from sentence_transformers import SentenceTransformer
    import torch

    device = "cuda" if torch.cuda.is_available() else "cpu"
    return SentenceTransformer(model_name, device=device)


def compute_embeddings(texts: list[str], provider: str, model_name: str) -> np.ndarray:
    import numpy as np

    if provider != "sentence-transformers":
        raise ValueError(f"Unsupported embedding provider: {provider}")
    model = load_sentence_transformer(model_name)
    batch_size = 32
    vectors = []
    total_batches = int(math.ceil(len(texts) / batch_size))
    for batch_index, start in enumerate(range(0, len(texts), batch_size), start=1):
        batch = texts[start : start + batch_size]
        encoded = model.encode(batch, batch_size=batch_size, show_progress_bar=False, normalize_embeddings=True)
        vectors.extend(encoded)
        print(
            json.dumps(
                {
                    "stage": "embedding_batch",
                    "batch": batch_index,
                    "total_batches": total_batches,
                    "encoded_rows": min(start + len(batch), len(texts)),
                },
                ensure_ascii=False,
            ),
            flush=True,
        )
    return np.asarray(vectors)


def nearest_examples(embeddings: np.ndarray, indices: list[int], centroid: np.ndarray, top_n: int) -> list[int]:
    import numpy as np

    vectors = embeddings[indices]
    scores = np.dot(vectors, centroid)
    ranked = [indices[i] for i in np.argsort(scores)[::-1]]
    return ranked[:top_n]


def choose_cluster_count(n_rows: int) -> int:
    return max(3, min(12, int(round(math.sqrt(max(9, n_rows)) / 1.6))))


def extract_json(text: str) -> Any:
    match = re.search(r"```json\s*(.+?)\s*```", text, re.S)
    if match:
        return json.loads(match.group(1))
    match = re.search(r"(\{.+\}|\[.+\])", text, re.S)
    if not match:
        raise ValueError("No JSON object found in model response")
    return json.loads(match.group(1))


def call_openai_json(model: str, api_key: str, prompt: str) -> Any:
    import requests

    payload = {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": [
                    {
                        "type": "input_text",
                        "text": "You are a strict taxonomy and dataset-analysis assistant. Return compact JSON only.",
                    }
                ],
            },
            {"role": "user", "content": [{"type": "input_text", "text": prompt}]},
        ],
        "text": {"format": {"type": "text"}},
        "max_output_tokens": 1200,
    }
    response = requests.post(
        "https://api.openai.com/v1/responses",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json=payload,
        timeout=120,
    )
    response.raise_for_status()
    data = response.json()
    text_parts = []
    for item in data.get("output", []):
        for content in item.get("content", []):
            if content.get("type") == "output_text":
                text_parts.append(content.get("text", ""))
    return extract_json("\n".join(text_parts))


def deterministic_cluster_summary(rows: list[TicketRow], indices: list[int]) -> dict[str, Any]:
    headers = rows[0].values.keys()
    candidate_fields = [
        find_column(headers, "unterkategorie"),
        find_column(headers, "kategorie"),
        find_column(headers, "art 'servicefall'", "art servicefall"),
        find_column(headers, "eingangsart"),
    ]
    values = []
    for field in candidate_fields:
        if not field:
            continue
        values.extend(str(rows[i].get(field) or "").strip() for i in indices if not is_blank(rows[i].get(field)))
    value_counts = Counter(value for value in values if value and value.lower() != "unknown")
    top_values = [value for value, _count in value_counts.most_common(3)]
    if top_values:
        label = " / ".join(top_values[:2])
    else:
        titles = []
        title_column = find_column(headers, "kurzbeschreibung", "details")
        for i in indices:
            text = str(rows[i].get(title_column) or "").strip()
            titles.extend(re.findall(r"[A-Za-zÄÖÜäöüß0-9._/-]{4,}", text))
        filtered = [token for token, _count in Counter(titles).most_common(4)]
        label = " / ".join(filtered[:2]) if filtered else "mixed operational pattern"
    definition = f"Recurring ticket pattern centered around {label}."
    return {
        "label": label,
        "definition": definition,
        "inclusion_signals": top_values[:3],
        "edge_case_note": "Deterministic label; refine with a small LLM later if needed.",
        "service_hints": top_values[:2],
        "noise": False,
    }


def build_semantic_taxonomy(
    rows: list[TicketRow],
    embedding_provider: str,
    embedding_model: str,
    openai_model: str,
    openai_key: str,
    naming_mode: str,
) -> dict[str, Any]:
    import numpy as np
    from sklearn.cluster import MiniBatchKMeans

    headers = rows[0].values.keys()
    texts = [build_semantic_text(row, headers) for row in rows]
    embeddings = compute_embeddings(texts, embedding_provider, embedding_model)
    cluster_count = choose_cluster_count(len(rows))
    model = MiniBatchKMeans(n_clusters=cluster_count, random_state=42, batch_size=256, n_init="auto")
    labels = model.fit_predict(embeddings)
    title_column = find_column(headers, "kurzbeschreibung", "details")
    buckets = []
    cluster_debug = []
    min_cluster_size = max(3, min(12, int(round(len(rows) / 8))))
    for cluster_id in range(cluster_count):
        indices = [i for i, label in enumerate(labels) if label == cluster_id]
        if len(indices) < min_cluster_size:
            continue
        centroid = model.cluster_centers_[cluster_id]
        representative_indices = nearest_examples(embeddings, indices, centroid, top_n=6)
        examples = [rows[i] for i in representative_indices]
        print(
            json.dumps(
                {"stage": "cluster_naming_start", "cluster_id": cluster_id, "cluster_size": len(indices)},
                ensure_ascii=False,
            ),
            flush=True,
        )
        prompt = {
            "task": "Name and define one semantic ticket-pattern cluster from example tickets.",
            "instructions": [
                "Return JSON with keys label, definition, inclusion_signals, edge_case_note, service_hints, noise.",
                "The label must sound like an operational taxonomy bucket, not a vague topic sentence.",
                "If the cluster is too mixed, set noise=true.",
            ],
            "examples": [
                {
                    "ticket_id": row.ticket_id,
                    "title": str(row.get(title_column) or row.ticket_id),
                    "text": shorten(build_semantic_text(row, headers), 700),
                }
                for row in examples
            ],
        }
        if naming_mode == "openai":
            result = call_openai_json(openai_model, openai_key, json.dumps(prompt, ensure_ascii=False))
        else:
            result = deterministic_cluster_summary(rows, representative_indices)
        cluster_debug.append({"cluster_id": cluster_id, "size": len(indices), "llm": result})
        if result.get("noise"):
            print(
                json.dumps({"stage": "cluster_naming_noise", "cluster_id": cluster_id}, ensure_ascii=False),
                flush=True,
            )
            continue
        bucket_name = result.get("label") or f"pattern_{cluster_id}"
        print(
            json.dumps(
                {"stage": "cluster_naming_done", "cluster_id": cluster_id, "bucket_name": bucket_name},
                ensure_ascii=False,
            ),
            flush=True,
        )
        buckets.append(
            {
                "bucket_name": bucket_name,
                "definition": result.get("definition", ""),
                "row_count_estimate": len(indices),
                "confidence": round(min(0.96, 0.65 + len(indices) / max(3000, len(rows) * 1.2)), 2),
                "canonical_examples": [example_for_row(rows[i], title_column) for i in representative_indices[:1]],
                "common_examples": [example_for_row(rows[i], title_column) for i in representative_indices[1:4]],
                "edge_examples": [example_for_row(rows[i], title_column) for i in representative_indices[4:5]],
                "inclusion_signals": result.get("inclusion_signals", []),
                "exclusion_signals": [],
                "service_hints": result.get("service_hints", []),
                "edge_case_note": result.get("edge_case_note"),
            }
        )
    return {
        "dimension_name": "issue_pattern",
        "purpose": "Semantic clustering of repeated issue shapes from ticket narrative and categorical context.",
        "inference_basis": "embedding_cluster_plus_llm_naming",
        "source_fields": [
            field
            for field in [
                find_column(headers, "kurzbeschreibung", "details"),
                find_column(headers, "unterkategorie"),
                find_column(headers, "kategorie"),
                find_column(headers, "name anmelder"),
                find_column(headers, "anfrage"),
                find_column(headers, "aktion"),
            ]
            if field
        ],
        "bucket_count": len(buckets),
        "stability_note": "Promoted from embedding clusters only where the small LLM judged the examples coherent enough to name.",
        "state": "promoted",
        "buckets": sorted(buckets, key=lambda b: b["row_count_estimate"], reverse=True),
        "cluster_debug": cluster_debug,
    }


def build_glossary(
    rows: list[TicketRow],
    structured_taxonomies: list[dict[str, Any]],
    semantic_taxonomy: dict[str, Any],
    openai_model: str,
    openai_key: str,
    glossary_mode: str,
) -> list[dict[str, Any]]:
    headers = rows[0].values.keys()
    text_fields = [
        field
        for field in [
            find_column(headers, "kurzbeschreibung", "details"),
            find_column(headers, "anfrage"),
            find_column(headers, "aktion"),
            find_column(headers, "unterkategorie"),
            find_column(headers, "kategorie"),
        ]
        if field
    ]

    token_counter: Counter[str] = Counter()
    for row in rows[: min(len(rows), 8000)]:
        for field in text_fields:
            value = row.get(field)
            if is_blank(value):
                continue
            text = str(value)
            for token in re.findall(r"\b[A-Z0-9][A-Z0-9._/-]{2,}\b", text):
                cleaned = token.strip(".,;:()[]{}<>")
                if len(cleaned) >= 3:
                    token_counter[cleaned] += 1

    structured_terms = set()
    for taxonomy in structured_taxonomies:
        for bucket in taxonomy.get("buckets", []):
            structured_terms.add(bucket["bucket_name"])
    semantic_terms = {bucket["bucket_name"] for bucket in semantic_taxonomy.get("buckets", [])}
    ranked_tokens = [token for token, count in token_counter.most_common(80) if count >= 3]

    candidates = list(dict.fromkeys(list(structured_terms)[:25] + list(semantic_terms)[:25] + ranked_tokens[:40]))[:60]
    prompt = {
        "task": "Clean glossary candidates for a service-desk knowledge base.",
        "instructions": [
            "Return JSON as a list of objects with keys term, meaning, kind, usage_note.",
            "Keep only stable technical, operational, service, platform, queue, or system terms.",
            "Discard generic German or English support words.",
            "meaning must be short and cautious when the exact semantics are unclear.",
        ],
        "candidates": candidates,
    }
    print(json.dumps({"stage": "glossary_start", "candidate_count": len(candidates)}, ensure_ascii=False), flush=True)
    if glossary_mode == "openai":
        result = call_openai_json(openai_model, openai_key, json.dumps(prompt, ensure_ascii=False))
        if not isinstance(result, list):
            raise ValueError("Glossary cleanup did not return a JSON list")
        glossary = []
        for item in result:
            term = str(item.get("term", "")).strip()
            if not term:
                continue
            glossary.append(
                {
                    "term": term,
                    "meaning": str(item.get("meaning", "")).strip(),
                    "kind": str(item.get("kind", "unknown")).strip(),
                    "usage_note": str(item.get("usage_note", "")).strip(),
                }
            )
    else:
        glossary = [
            {
                "term": term,
                "meaning": "Observed repeated operational or system term in the ticket dataset.",
                "kind": "observed_term",
                "usage_note": "Deterministic glossary entry; refine with a small LLM later if needed.",
            }
            for term in candidates[:25]
        ]
    return glossary


def projection_guidance(
    structured_taxonomies: list[dict[str, Any]],
    semantic_taxonomy: dict[str, Any],
    glossary: list[dict[str, Any]],
) -> dict[str, Any]:
    dimension_names = {taxonomy["dimension_name"] for taxonomy in structured_taxonomies}
    semantic_bucket_names = [bucket["bucket_name"] for bucket in semantic_taxonomy.get("buckets", [])]
    glossary_terms = [entry["term"] for entry in glossary[:20]]
    return {
        "ticket_onboarding": {
            "label_candidate_sources": [
                name
                for name in ["category_tree", "service_category", "subcategory", "request_type", "issue_pattern"]
                if name in dimension_names or name == semantic_taxonomy.get("dimension_name")
            ],
            "ownership_hint_sources": [name for name in ["reporter_family", "site_scope", "workflow_state"] if name in dimension_names],
            "initial_review_focus": semantic_bucket_names[:8],
        },
        "service_mapping": {
            "strong_dimensions": [name for name in ["service_category", "category_tree", "issue_pattern"] if name in dimension_names or name == "issue_pattern"],
            "supporting_terms": glossary_terms,
        },
        "monitoring_and_access": {
            "candidate_signals": [
                bucket["bucket_name"]
                for bucket in semantic_taxonomy.get("buckets", [])
                if any(
                    needle in bucket["bucket_name"].lower()
                    for needle in ["monitor", "alert", "zugang", "account", "mail", "ftp", "druck", "server", "login"]
                )
            ][:10],
            "note": "These buckets should be cross-checked against monitoring and access knowledge before granting autonomy.",
        },
    }


def render_knowledgebase_markdown(
    source_profile: dict[str, Any],
    structured_taxonomies: list[dict[str, Any]],
    semantic_taxonomy: dict[str, Any],
    glossary: list[dict[str, Any]],
    guidance: dict[str, Any],
) -> str:
    lines = []
    lines.append("# Ticket Dataset Knowledge Base")
    lines.append("")
    scope = source_profile["source_scope"]
    lines.append("## Source Profile")
    lines.append("")
    lines.append(f"- Sheet: `{scope['sheet_name']}`")
    lines.append(f"- Rows: `{scope['row_count']}`")
    lines.append(f"- Columns: `{scope['column_count']}`")
    lines.append(f"- Row unit: {source_profile['row_unit_description']}")
    lines.append("")
    lines.append("## Promoted Structured Taxonomies")
    lines.append("")
    for taxonomy in structured_taxonomies:
        lines.append(f"### {taxonomy['dimension_name']}")
        lines.append("")
        lines.append(f"- Purpose: {taxonomy['purpose']}")
        lines.append(f"- Source fields: {', '.join(f'`{field}`' for field in taxonomy['source_fields'])}")
        lines.append(f"- Buckets: `{taxonomy['bucket_count']}`")
        lines.append("")
        for bucket in taxonomy["buckets"][:6]:
            examples = ", ".join(f"{ex['ticket_id']} {ex['title']}" for ex in bucket["canonical_examples"][:1])
            lines.append(
                f"- `{bucket['bucket_name']}`: {bucket['definition']} "
                f"(rows≈{bucket['row_count_estimate']}, example: {examples})"
            )
        lines.append("")

    lines.append("## Semantic Issue-Pattern Taxonomy")
    lines.append("")
    lines.append(f"- Buckets: `{semantic_taxonomy['bucket_count']}`")
    lines.append(f"- Source fields: {', '.join(f'`{field}`' for field in semantic_taxonomy['source_fields'])}")
    lines.append("")
    for bucket in semantic_taxonomy["buckets"][:10]:
        example_ids = ", ".join(ex["ticket_id"] for ex in bucket["canonical_examples"][:1] + bucket["common_examples"][:2])
        lines.append(f"### {bucket['bucket_name']}")
        lines.append("")
        lines.append(f"- Definition: {bucket['definition']}")
        lines.append(f"- Estimated rows: `{bucket['row_count_estimate']}`")
        if bucket.get("service_hints"):
            lines.append(f"- Service hints: {', '.join(bucket['service_hints'])}")
        if bucket.get("inclusion_signals"):
            lines.append(f"- Inclusion signals: {', '.join(bucket['inclusion_signals'][:5])}")
        lines.append(f"- Examples: {example_ids}")
        if bucket.get("edge_case_note"):
            lines.append(f"- Edge note: {bucket['edge_case_note']}")
        lines.append("")

    lines.append("## Glossary")
    lines.append("")
    for entry in glossary[:25]:
        suffix = f" ({entry['kind']})" if entry.get("kind") else ""
        usage = f" Usage: {entry['usage_note']}" if entry.get("usage_note") else ""
        lines.append(f"- `{entry['term']}`{suffix}: {entry['meaning']}.{usage}")
    lines.append("")

    lines.append("## Projection Guidance")
    lines.append("")
    ticket_onboarding = guidance["ticket_onboarding"]
    lines.append(
        f"- Ticket onboarding label candidates: {', '.join(ticket_onboarding['label_candidate_sources']) or 'none'}"
    )
    lines.append(f"- Ownership hint sources: {', '.join(ticket_onboarding['ownership_hint_sources']) or 'none'}")
    lines.append(
        f"- First review focus: {', '.join(ticket_onboarding['initial_review_focus']) or 'none'}"
    )
    lines.append("")
    lines.append("## Why This Counts As A Knowledge Base")
    lines.append("")
    lines.append("- The source profile captures what one row means and how the table is structured.")
    lines.append("- Structured taxonomies promote repeatable operational dimensions from durable columns.")
    lines.append("- Semantic issue-pattern buckets preserve recurring problem shapes beyond flat labels.")
    lines.append("- Each bucket carries representative examples so later ticket work can anchor on evidence.")
    lines.append("- The glossary stabilizes repeated terms, systems, and operational language.")
    lines.append("- Projection guidance explains how later onboarding and ticket work should consume the knowledge.")
    lines.append("")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a ticket dataset knowledge base from an XLSX export.")
    parser.add_argument("--input-xlsx", required=True, help="Path to the input XLSX workbook.")
    parser.add_argument("--output-dir", required=True, help="Directory where knowledge artifacts will be written.")
    parser.add_argument(
        "--embedding-provider",
        default="sentence-transformers",
        help="Embedding provider. Currently only sentence-transformers is supported.",
    )
    parser.add_argument(
        "--embedding-model",
        default="Qwen/Qwen3-Embedding-0.6B",
        help="Embedding model name for sentence-transformers.",
    )
    parser.add_argument("--openai-model", default="gpt-5.4-nano", help="OpenAI model for taxonomy naming.")
    parser.add_argument(
        "--openai-api-key-env",
        default="OPENAI_API_KEY",
        help="Environment variable that contains the OpenAI API key.",
    )
    parser.add_argument(
        "--max-semantic-rows",
        type=int,
        default=4000,
        help="Maximum number of rows to use for semantic clustering.",
    )
    parser.add_argument(
        "--semantic-naming-mode",
        choices=["openai", "deterministic"],
        default="openai",
        help="How semantic clusters are named after embeddings.",
    )
    parser.add_argument(
        "--glossary-mode",
        choices=["openai", "deterministic"],
        default="openai",
        help="How glossary entries are cleaned and defined.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_xlsx)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    openai_key = os.environ.get(args.openai_api_key_env)
    if not openai_key:
        raise RuntimeError(f"Missing OpenAI API key in environment variable {args.openai_api_key_env}")

    sheet_name, rows = load_ticket_rows(input_path)
    if not rows:
        raise RuntimeError("The workbook did not contain any non-empty ticket rows")
    print(json.dumps({"stage": "load", "rows": len(rows), "sheet_name": sheet_name}, ensure_ascii=False), flush=True)

    source_profile = build_source_profile(sheet_name, rows)
    print(
        json.dumps(
            {
                "stage": "source_profile",
                "key_columns": source_profile["key_columns"],
                "enum_like_columns": len(source_profile["enum_like_columns"]),
                "free_text_columns": len(source_profile["free_text_columns"]),
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    structured_taxonomies = promote_structured_taxonomies(rows)
    print(
        json.dumps(
            {
                "stage": "structured_taxonomies",
                "count": len(structured_taxonomies),
                "dimensions": [taxonomy["dimension_name"] for taxonomy in structured_taxonomies],
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    semantic_rows = sample_rows_for_semantics(rows, args.max_semantic_rows)
    print(
        json.dumps(
            {"stage": "semantic_sampling", "semantic_row_count": len(semantic_rows), "max_semantic_rows": args.max_semantic_rows},
            ensure_ascii=False,
        ),
        flush=True,
    )
    semantic_taxonomy = build_semantic_taxonomy(
        semantic_rows,
        args.embedding_provider,
        args.embedding_model,
        args.openai_model,
        openai_key,
        args.semantic_naming_mode,
    )
    print(
        json.dumps(
            {"stage": "semantic_taxonomy", "bucket_count": semantic_taxonomy["bucket_count"]},
            ensure_ascii=False,
        ),
        flush=True,
    )
    glossary = build_glossary(
        rows,
        structured_taxonomies,
        semantic_taxonomy,
        args.openai_model,
        openai_key,
        args.glossary_mode,
    )
    print(json.dumps({"stage": "glossary", "count": len(glossary)}, ensure_ascii=False), flush=True)
    guidance = projection_guidance(structured_taxonomies, semantic_taxonomy, glossary)
    knowledgebase_md = render_knowledgebase_markdown(
        source_profile,
        structured_taxonomies,
        semantic_taxonomy,
        glossary,
        guidance,
    )

    outputs = {
        "source_profile.json": source_profile,
        "taxonomies.json": {"structured": structured_taxonomies, "semantic": semantic_taxonomy},
        "semantic_clusters.json": semantic_taxonomy,
        "glossary.json": glossary,
        "projection_guidance.json": guidance,
    }
    for filename, payload in outputs.items():
        (output_dir / filename).write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, default=json_default) + "\n",
            encoding="utf-8",
        )
    (output_dir / "knowledgebase.md").write_text(knowledgebase_md + "\n", encoding="utf-8")
    print(json.dumps({"stage": "write_outputs", "output_dir": str(output_dir)}, ensure_ascii=False), flush=True)

    summary = {
        "input_xlsx": str(input_path),
        "output_dir": str(output_dir),
        "row_count": len(rows),
        "semantic_row_count": len(semantic_rows),
        "structured_taxonomy_count": len(structured_taxonomies),
        "semantic_bucket_count": semantic_taxonomy["bucket_count"],
        "glossary_count": len(glossary),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
